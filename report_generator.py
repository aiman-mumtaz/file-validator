import csv
import io
import logging
import re
from datetime import datetime
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

OVERPUNCH_POSITIVE = {
    '{': '0', 'A': '1', 'B': '2', 'C': '3', 'D': '4',
    'E': '5', 'F': '6', 'G': '7', 'H': '8', 'I': '9'
}
OVERPUNCH_NEGATIVE = {
    '}': '0', 'J': '1', 'K': '2', 'L': '3', 'M': '4',
    'N': '5', 'O': '6', 'P': '7', 'Q': '8', 'R': '9'
}


class ValidationError(Exception):
    pass


def _normalize_header(value: str) -> str:
    return re.sub(r'[^a-z0-9]', '', value.strip().lower()) if value else ''


def _parse_bool(value: str) -> bool:
    if value is None:
        return False
    return str(value).strip().lower() in {'true', 'yes', 'y', '1'}


def _parse_master_seq(value: str) -> int | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned:
        return None
    try:
        return int(float(cleaned))
    except ValueError:
        return None


def _load_field_mapping(mapping_stream: io.TextIOBase) -> Dict[str, dict]:
    field_mapping: Dict[str, dict] = {}
    reader = csv.DictReader(mapping_stream)
    header_map = { _normalize_header(name): name for name in (reader.fieldnames or []) }
    logger.info("Mapping headers detected: %s", ", ".join(reader.fieldnames or []))

    def get_field(row: dict, *candidates: str) -> str | None:
        for candidate in candidates:
            key = header_map.get(_normalize_header(candidate))
            if key and key in row:
                return row.get(key)
        for candidate in candidates:
            if candidate in row:
                return row.get(candidate)
        return None

    for row in reader:
        try:
            master_seq_val = _parse_master_seq(get_field(row, 'Master Seq No', 'MasterSeqNo', 'Master Sequence No'))
            if master_seq_val != 4:
                continue

            field_name = get_field(row, 'Field Name', 'FieldName')
            field_format = get_field(row, 'Field Format', 'FieldFormat')
            field_length_raw = get_field(row, 'Field Length', 'FieldLength')
            field_length = int(field_length_raw) if field_length_raw else 0

            is_numerical = _parse_bool(get_field(row, 'IsNumericalFormat', 'IsNumerical'))
            is_overpunch = _parse_bool(get_field(row, 'IsOverpunchFormat', 'IsOverpunch'))
            format_spec = get_field(row, 'Format', 'FieldFormatSpec')
            if format_spec and str(format_spec).strip().lower() == 'null':
                format_spec = None

            start_pos = get_field(row, 'start', 'Start', 'Start Position', 'StartPosition')
            end_pos = get_field(row, 'end', 'End', 'End Position', 'EndPosition')

            apply_alpha_pad = get_field(row, 'applyAlphaPad', 'ApplyAlphaPad') or ''
            apply_alpha_len = get_field(row, 'applyAlphaLen', 'ApplyAlphaLen') or ''
            apply_numeric_pad = get_field(row, 'applyNumericPad', 'ApplyNumericPad') or ''
            apply_numeric_len = get_field(row, 'applyNumericLen', 'ApplyNumericLen') or ''

            if field_length > 0 and field_name and start_pos and end_pos:
                decimal_places = 0
                if format_spec and 'V' in str(format_spec):
                    match = re.search(r'V9+', str(format_spec))
                    if match:
                        decimal_places = len(match.group()) - 1

                alpha_pad_len = int(apply_alpha_len) if str(apply_alpha_len).isdigit() else 0
                numeric_pad_len = int(apply_numeric_len) if str(apply_numeric_len).isdigit() else 0

                field_mapping[str(field_name).strip()] = {
                    'format': field_format,
                    'length': field_length,
                    'is_numerical': is_numerical,
                    'is_overpunch': is_overpunch,
                    'decimal_places': decimal_places,
                    'format_spec': format_spec,
                    'start': int(float(start_pos)) - 1,
                    'end': int(float(end_pos)),
                    'alpha_pad_field': apply_alpha_pad,
                    'alpha_pad_len': alpha_pad_len,
                    'numeric_pad_field': apply_numeric_pad,
                    'numeric_pad_len': numeric_pad_len
                }
        except (ValueError, KeyError, TypeError):
            continue

    if not field_mapping:
        found_headers = ', '.join(reader.fieldnames or [])
        logger.error("No field mappings found in CSV. Headers: %s", found_headers or 'none')
        raise ValidationError(
            "No field mappings found in CSV. Verify the mapping file format. "
            f"Found headers: {found_headers or 'none'}"
        )

    return field_mapping


def decode_overpunch(value: str) -> str:
    if not value:
        return value

    last_char = value[-1]
    if last_char in OVERPUNCH_POSITIVE:
        return value[:-1] + OVERPUNCH_POSITIVE[last_char]
    if last_char in OVERPUNCH_NEGATIVE:
        return '-' + value[:-1] + OVERPUNCH_NEGATIVE[last_char]
    return value


def normalize_field(raw_value: str, field_info: dict) -> str:
    if not raw_value:
        return ''

    cleaned = ''.join(char if char.isprintable() else ' ' for char in raw_value).strip()

    if not cleaned:
        return ''

    is_numerical = field_info.get('is_numerical', False)
    is_overpunch = field_info.get('is_overpunch', False)
    decimal_places = field_info.get('decimal_places', 0)

    if is_overpunch:
        cleaned = decode_overpunch(cleaned)

    if is_numerical:
        try:
            num_str = ''.join(c for c in cleaned if c.isdigit() or c == '-')
            if not num_str or num_str == '-' or num_str.replace('-', '') == '0' * len(num_str.replace('-', '')):
                return ''

            is_negative = num_str.startswith('-')
            num_str = num_str.replace('-', '')

            if decimal_places > 0 and num_str:
                if len(num_str) > decimal_places:
                    integer_part = num_str[:-decimal_places]
                    fractional_part = num_str[-decimal_places:]
                    num_val = float(f"{integer_part}.{fractional_part}")
                else:
                    num_val = float(f"0.{num_str.zfill(decimal_places)}")

                if is_negative:
                    num_val = -num_val

                result = str(num_val)
                if '.' in result:
                    result = result.rstrip('0').rstrip('.')
                return result

            num_val = int(num_str)
            if is_negative:
                num_val = -num_val
            return str(num_val)
        except Exception:
            pass

    return cleaned


def parse_claim_with_mapping(line: str, field_mapping: Dict[str, dict]) -> dict:
    if not line.startswith('4'):
        return None

    claim = {}
    claim_raw = {}
    claim_raw_exact = {}

    for field_name, field_info in field_mapping.items():
        start = field_info['start']
        end = field_info['end']
        raw_value = line[start:end]
        cleaned_raw = ''.join(char if char.isprintable() else ' ' for char in raw_value).strip()
        normalized_value = normalize_field(raw_value, field_info)
        claim[field_name] = normalized_value
        claim_raw[field_name] = cleaned_raw
        claim_raw_exact[field_name] = raw_value

    claim['_raw'] = claim_raw
    claim['_raw_exact'] = claim_raw_exact
    return claim


def create_record_signature(claim: dict) -> Tuple[str, ...]:
    return (
        claim.get('RX CLAIMS NUMBER', ''),
        claim.get('CLAIM STATUS', ''),
        claim.get('SEQUENCE NUMBER OF CLAIM', ''),
        claim.get('PATIENT FIRST NAME', ''),
        claim.get('PATIENT LAST NAME', ''),
        claim.get('PATIENT DATE OF BIRTH', ''),
    )


def _read_claims(lines: Iterable[str], field_mapping: Dict[str, dict]) -> List[dict]:
    claims = []
    for line_num, line in enumerate(lines, 1):
        if line.startswith('4'):
            claim = parse_claim_with_mapping(line, field_mapping)
            if claim:
                claim['Line_Number'] = line_num
                claims.append(claim)
    return claims


def _compare_claims(rxe_claims: List[dict], adhoc_claims: List[dict], field_mapping: Dict[str, dict]) -> Tuple[List[dict], List[Tuple[int, dict]], List[Tuple[int, dict]], dict, List[str]]:
    rxe_lookup = {create_record_signature(claim): (idx, claim) for idx, claim in enumerate(rxe_claims)}
    adhoc_lookup = {create_record_signature(claim): (idx, claim) for idx, claim in enumerate(adhoc_claims)}

    value_differences = []
    missing_in_adhoc = []
    extra_in_adhoc = []

    signature_fields = {
        'RX CLAIMS NUMBER',
        'CLAIM STATUS',
        'SEQUENCE NUMBER OF CLAIM',
        'PATIENT FIRST NAME',
        'PATIENT LAST NAME',
        'PATIENT DATE OF BIRTH',
    }

    compare_fields = [field for field in field_mapping.keys() if field not in signature_fields]

    for signature in rxe_lookup:
        rxe_idx, rxe_claim = rxe_lookup[signature]

        if signature in adhoc_lookup:
            _, adhoc_claim = adhoc_lookup[signature]

            for field in compare_fields:
                rxe_val = rxe_claim.get(field, '')
                adhoc_val = adhoc_claim.get(field, '')

                if rxe_val != adhoc_val:
                    rxe_exact = rxe_claim.get('_raw_exact', {}).get(field, '')
                    adhoc_exact = adhoc_claim.get('_raw_exact', {}).get(field, '')
                    if rxe_exact == adhoc_exact:
                        continue

                    rxe_raw = rxe_claim.get('_raw', {}).get(field, rxe_val)
                    adhoc_raw = adhoc_claim.get('_raw', {}).get(field, adhoc_val)

                    value_differences.append({
                        'Rx_Number': rxe_claim.get('PRESCRIPTION /SERVICE REFERENCE NUMBER', ''),
                        'Rx_Claims_Number': rxe_claim.get('RX CLAIMS NUMBER', ''),
                        'Claim_Status': rxe_claim.get('CLAIM STATUS', ''),
                        'Sequence_Number_Of_Claim': rxe_claim.get('SEQUENCE NUMBER OF CLAIM', ''),
                        'Patient_First': rxe_claim.get('PATIENT FIRST NAME', ''),
                        'Patient_Last': rxe_claim.get('PATIENT LAST NAME', ''),
                        'Patient_DOB': rxe_claim.get('PATIENT DATE OF BIRTH', ''),
                        'Date_Of_Service': rxe_claim.get('DATE OF SERVICE', ''),
                        'Drug': rxe_claim.get('PRODUCT LABEL NAME WITH DOSAGE FORM AND STRENGTH', ''),
                        'Field': field,
                        'Correct_RxE': rxe_raw,
                        'Wrong_ADHOC': adhoc_raw
                    })
        else:
            missing_in_adhoc.append((rxe_idx + 1, rxe_claim))

    for signature in adhoc_lookup:
        if signature not in rxe_lookup:
            adhoc_idx, adhoc_claim = adhoc_lookup[signature]
            extra_in_adhoc.append((adhoc_idx + 1, adhoc_claim))

    differences_by_field = {}
    for diff in value_differences:
        field = diff['Field']
        differences_by_field.setdefault(field, []).append(diff)

    return value_differences, missing_in_adhoc, extra_in_adhoc, differences_by_field, compare_fields


def generate_excel_report(
    mapping_stream: io.TextIOBase,
    rxe_stream: io.TextIOBase,
    adhoc_stream: io.TextIOBase,
    timestamp: str | None = None
) -> Tuple[bytes, dict]:
    logger.info("Generating validation report")
    field_mapping = _load_field_mapping(mapping_stream)
    rxe_claims = _read_claims(rxe_stream, field_mapping)
    adhoc_claims = _read_claims(adhoc_stream, field_mapping)
    logger.info("Parsed %s RxE claims and %s SSE claims", len(rxe_claims), len(adhoc_claims))

    (
        value_differences,
        missing_in_adhoc,
        extra_in_adhoc,
        differences_by_field,
        _
    ) = _compare_claims(rxe_claims, adhoc_claims, field_mapping)

    wb = Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    highlight_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary['A1'] = 'SSE CLAIMS VALIDATION REPORT'
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')

    ws_summary['A2'] = 'CHF 7.0 Format - RxE as Reference'
    ws_summary['A2'].font = Font(italic=True, size=10)
    ws_summary.merge_cells('A2:B2')

    ws_summary['A4'] = 'Metric'
    ws_summary['B4'] = 'Count'
    ws_summary['A4'].fill = header_fill
    ws_summary['B4'].fill = header_fill
    ws_summary['A4'].font = header_font
    ws_summary['B4'].font = header_font

    summary_data = [
        ['Total Claims in RxE_Test (Reference)', len(rxe_claims)],
        ['Total Claims in ADHOC (Validation)', len(adhoc_claims)],
        ['', ''],
        ['Total Field Value Differences', len(value_differences)],
        ['Number of Different Fields', len(differences_by_field)],
        ['Missing Claims in ADHOC', len(missing_in_adhoc)],
        ['Extra Claims in ADHOC', len(extra_in_adhoc)],
    ]

    row = 5
    for item in summary_data:
        ws_summary[f'A{row}'] = item[0]
        ws_summary[f'B{row}'] = item[1]
        if item[0] and not item[1] == '':
            ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1

    if differences_by_field:
        ws_summary[f'A{row + 1}'] = 'Differences by Field Type (Click worksheet tabs below)'
        ws_summary[f'A{row + 1}'].font = Font(bold=True, size=11, color="4472C4")
        ws_summary.merge_cells(f'A{row + 1}:B{row + 1}')

        ws_summary[f'A{row + 3}'] = 'Field Name'
        ws_summary[f'B{row + 3}'] = 'Number of Differences'
        ws_summary[f'A{row + 3}'].fill = header_fill
        ws_summary[f'B{row + 3}'].fill = header_fill
        ws_summary[f'A{row + 3}'].font = header_font
        ws_summary[f'B{row + 3}'].font = header_font

        row += 4
        for field, count in sorted(differences_by_field.items(), key=lambda x: len(x[1]), reverse=True):
            ws_summary[f'A{row}'] = field
            ws_summary[f'B{row}'] = len(count)
            row += 1

    ws_summary.column_dimensions['A'].width = 60
    ws_summary.column_dimensions['B'].width = 25

    for field_name, diffs in sorted(differences_by_field.items(), key=lambda x: len(x[1]), reverse=True):
        sheet_name = field_name[:30].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace('[', '').replace(']', '')
        ws_field = wb.create_sheet(sheet_name)

        ws_field['A1'] = f'Field: {field_name}'
        ws_field['A1'].font = Font(bold=True, size=13)
        ws_field.merge_cells('A1:I1')

        ws_field['A2'] = f'Total Differences: {len(diffs)}'
        ws_field['A2'].font = Font(bold=True, size=11, color="C00000")
        ws_field.merge_cells('A2:I2')

        headers = [
            'Rx Claims Number', 'Claim Status', 'Sequence Number of Claim',
            'Patient First Name', 'Patient Last Name', 'Patient Date of Birth',
            'Date of Service', 'Correct Value (RxE)', 'Incorrect Value (ADHOC)'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_field.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row = 5
        for diff in diffs:
            cell_a = ws_field.cell(row=row, column=1, value=diff['Rx_Claims_Number'])
            cell_a.number_format = '@'

            cell_b = ws_field.cell(row=row, column=2, value=diff.get('Claim_Status', ''))
            cell_b.number_format = '@'

            cell_c = ws_field.cell(row=row, column=3, value=diff.get('Sequence_Number_Of_Claim', ''))
            cell_c.number_format = '@'

            ws_field.cell(row=row, column=4, value=diff['Patient_First'])
            ws_field.cell(row=row, column=5, value=diff['Patient_Last'])
            ws_field.cell(row=row, column=6, value=diff.get('Patient_DOB', ''))

            cell_h = ws_field.cell(row=row, column=7, value=diff.get('Date_Of_Service', ''))
            cell_h.number_format = '@'

            rxe_cell = ws_field.cell(row=row, column=8, value=diff['Correct_RxE'])
            rxe_cell.fill = highlight_fill
            rxe_cell.number_format = '@'

            adhoc_cell = ws_field.cell(row=row, column=9, value=diff['Wrong_ADHOC'])
            adhoc_cell.fill = highlight_fill
            adhoc_cell.number_format = '@'

            row += 1

        ws_field.column_dimensions['A'].width = 20
        ws_field.column_dimensions['B'].width = 16
        ws_field.column_dimensions['C'].width = 20
        ws_field.column_dimensions['D'].width = 16
        ws_field.column_dimensions['E'].width = 16
        ws_field.column_dimensions['F'].width = 14
        ws_field.column_dimensions['G'].width = 18
        ws_field.column_dimensions['H'].width = 25
        ws_field.column_dimensions['I'].width = 25

        if row > 5:
            ws_field.auto_filter.ref = f"A4:I{row - 1}"

        ws_field.freeze_panes = 'A5'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if not timestamp:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    summary = {
        'timestamp': timestamp,
        'total_rxe': len(rxe_claims),
        'total_sse': len(adhoc_claims),
        'total_differences': len(value_differences),
        'fields_with_differences': len(differences_by_field),
        'missing_in_sse': len(missing_in_adhoc),
        'extra_in_sse': len(extra_in_adhoc),
    }

    return output.getvalue(), summary
